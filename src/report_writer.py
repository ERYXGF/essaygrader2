"""Writes the Excel grading report.
 
The report has up to three sheets:
- Summary    : one row per candidate, headline grades, colour-coded classification,
               a Double Application column marking candidates who applied for
               more than one role,
               an Embargo column flagging candidates who re-applied within six
               months of an earlier campaign's application,
               a File Format column flagging submissions that aren't PDFs,
               a Rubric Version column (shaded when the row was graded under an
               older rubric than this run, as happens after a role-scoped
               regrade),
               three feedback columns (Human Override / Override Reason /
               Reviewed) ready for the eventual RAG phase, and a Plagiarism Flag
               column (colour-coded, empty when clean).
- Detailed   : strengths, weaknesses, rationale, AI indicators, and one
               summary column per question. The Q1/Q2/Q3 columns are filled by
               canonical question_number, not by position in the list, so a
               candidate who answered out of sequence still lines up with
               everyone else.
- Similarity : one row per flagged essay pair with the full plagiarism evidence
               (only written when similarity_pairs is provided).
"""

from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def write_report(
    results: List[Dict],
    output_path: str,
    similarity_pairs: Optional[List[Dict]] = None,
) -> None:
    """Writes the grading results to an .xlsx file at output_path.

    similarity_pairs is the output of plagiarism_checker.check_plagiarism().
    Pass None (the default) to skip the Similarity sheet entirely — existing
    callers are unaffected. An empty list writes the sheet with a
    "no pairs flagged" row, so a clean run is distinguishable from no run.
    """
    if not results:
        raise ValueError("No results provided to write report")

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # =========================
    # STYLES
    # =========================
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ============================================================
    # SHEET 1 — SUMMARY
    # ============================================================
    # The last three columns are intentionally left empty on first write.
    # The reviewer fills them in; a future feedback_capture script will
    # ingest completed rows into the RAG database.
    ws1 = wb.active
    ws1.title = "Summary"

    double_applicants = _double_application_numbers(results)

    summary_headers = [
        "Candidate Number",
        "Role",
        "Double Application",
        "Embargo",
        "Classification",
        "Passion",
        "Humility",
        "Potential",
        "Writing Quality",
        "AI Risk",
        "File Format",
        "Rubric Version",
        "Human Override",
        "Override Reason",
        "Reviewed",
        "Plagiarism Flag",
    ]

    for col, h in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = center

    for row_idx, r in enumerate(results, start=2):
        cca = r.get("cross_cutting_assessment", {}) or {}
        writing = r.get("writing_quality", {}) or {}

        values = [
            r.get("candidate_number", "Unknown"),
            r.get("Role", "Unknown"),
            "Yes" if r.get("candidate_number") in double_applicants else "No",
            r.get("embargo", ""),
            r.get("classification", "Unknown"),
            cca.get("passion", ""),
            cca.get("humility", ""),
            cca.get("potential", ""),
            writing.get("rating", ""),
            r.get("ai_usage_probability", ""),
            r.get("format_label", ""),
            r.get("rubric_version", ""),
            "",  # Human Override (blank — reviewer fills in)
            "",  # Override Reason
            "",  # Reviewed
            r.get("plagiarism_flag", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.alignment = center

        # File format: a readable Word/Pages submission is graded normally,
        # but the interviewers cannot open it, so highlight the row.
        format_cell = ws1.cell(row=row_idx, column=summary_headers.index("File Format") + 1)
        if str(format_cell.value or "").startswith("wrong format"):
            format_cell.fill = yellow

        # A role-scoped regrade leaves other roles on an older rubric. Mixing
        # is acceptable, but it must be visible — a reviewer comparing two rows
        # needs to know they were judged against different criteria.
        rubric_cell = ws1.cell(row=row_idx, column=summary_headers.index("Rubric Version") + 1)
        if r.get("rubric_is_current") is False:
            rubric_cell.fill = yellow

        # Embargo: a candidate re-applying inside the six-month window. Long
        # text, so it wraps; red because the policy says they should not be in
        # this campaign at all. A blank cell means "checked and clear"; the
        # unknown marker means the recruitment list had nothing on them.
        embargo_cell = ws1.cell(
            row=row_idx, column=summary_headers.index("Embargo") + 1
        )
        embargo_cell.alignment = Alignment(vertical="center", wrap_text=True)
        embargo_value = str(embargo_cell.value or "")
        if embargo_value.startswith("⚠"):
            embargo_cell.fill = red
        elif embargo_value:
            embargo_cell.fill = yellow  # unknown — not cleared, not condemned

        # Plagiarism flag: wrap (can hold one line per matched pair) and
        # colour-band by risk so flagged rows stand out at a glance.
        flag_cell = ws1.cell(row=row_idx, column=len(summary_headers))
        flag_cell.alignment = Alignment(vertical="center", wrap_text=True)
        flag_value = str(flag_cell.value or "")
        if "⚠ HIGH" in flag_value:
            flag_cell.fill = red
        elif "⚠ MEDIUM" in flag_value:
            flag_cell.fill = yellow

        # Classification colour band. Located by header name, not position —
        # inserting a column ahead of it would otherwise silently colour the
        # wrong cells.
        class_cell = ws1.cell(
            row=row_idx, column=summary_headers.index("Classification") + 1
        )
        classification = r.get("classification", "")

        if classification == "Priority Interview":
            class_cell.fill = green
        elif classification == "Maybe":
            class_cell.fill = yellow
        elif "Do Not Interview" in classification:
            class_cell.fill = red

    _autosize(ws1, max_width=40)

    # ============================================================
    # SHEET 2 — DETAILED
    # ============================================================
    ws2 = wb.create_sheet(title="Detailed")

    detail_headers = [
        "Candidate Number",
        "Source File",
        "Strengths",
        "Weaknesses",
        "Rationale",
        "AI Indicators",
        "Q1 Summary",
        "Q2 Summary",
        "Q3 Summary",
    ]

    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font

    for row_idx, r in enumerate(results, start=2):
        strengths = r.get("strengths", []) or []
        weaknesses = r.get("weaknesses", []) or []
        summaries = _question_summaries(r.get("question_assessments", []) or [])

        values = [
            r.get("candidate_number", "Unknown"),
            r.get("source_file", ""),
            ", ".join(strengths),
            ", ".join(weaknesses),
            r.get("rationale", ""),
            r.get("ai_usage_indicators", ""),
            summaries.get(1, ""),
            summaries.get(2, ""),
            summaries.get(3, ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            if col >= 3:  # every long-form column
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autosize(ws2, max_width=60)

    # ============================================================
    # SHEET 3 — SIMILARITY (one row per flagged essay pair)
    # ============================================================
    if similarity_pairs is not None:
        _write_similarity_sheet(wb, similarity_pairs, header_font, red, yellow)

    # =========================
    # SAVE
    # =========================
    wb.save(output_file)


def _double_application_numbers(results: List[Dict]) -> set:
    """Candidate numbers appearing on more than one row.

    A candidate may legitimately apply for several roles — uniqueness is
    (candidate_number, role), so the same number showing twice is one person
    applying twice, not two people. Nothing in the report said so, which is
    what this answers. Rows with no candidate number are ignored rather than
    being grouped together under a shared blank.
    """
    seen, repeated = set(), set()
    for result in results:
        number = result.get("candidate_number")
        if not number:
            continue
        if number in seen:
            repeated.add(number)
        seen.add(number)
    return repeated


def _question_summaries(assessments: List[Dict]) -> Dict[int, str]:
    """Maps canonical question number -> summary text.

    Candidates answer the questions in whatever order they like, so the model
    is asked to identify each answer by content and report the canonical
    question number. Keying on that number — rather than on position in the
    list — is what keeps the Q1 column meaning the same question on every row.

    Results graded before question numbers were canonical (and any where the
    model omitted the field) carry no usable number, so those fall back to
    list order. TFO submissions have only two questions and simply leave Q3
    empty.
    """
    summaries: Dict[int, str] = {}
    unnumbered = []

    for entry in assessments:
        if not isinstance(entry, dict):
            continue
        number = entry.get("question_number")
        # A number we can't read is treated as absent, not as an error.
        try:
            number = int(number)
        except (TypeError, ValueError):
            number = None

        if number is None:
            unnumbered.append(entry)
        elif number in (1, 2, 3):
            summaries.setdefault(number, entry.get("summary", "") or "")
        # An out-of-range number (Q9, Q0) is a claim we can't honour. Drop it
        # rather than guessing a column — a blank cell is recoverable, the
        # wrong answer under the Q1 header is the bug this mapping exists to
        # prevent.

    # Fall back to position only for the columns nothing claimed by number.
    for position, entry in enumerate(unnumbered, start=1):
        if position > 3:
            break
        summaries.setdefault(position, entry.get("summary", "") or "")

    return summaries


def _write_similarity_sheet(
    wb: Workbook,
    pairs: List[Dict],
    header_font: Font,
    red: PatternFill,
    yellow: PatternFill,
) -> None:
    """Writes the per-pair plagiarism evidence sheet.

    Pairs come pre-ordered from plagiarism_checker (lower candidate number
    first); each row carries the two lexical/semantic screen scores, the
    risk band, and Claude's verdict with quoted evidence.
    """
    ws = wb.create_sheet(title="Similarity")

    headers = [
        "Candidate A",
        "Candidate B",
        "Lexical %",
        "Semantic %",
        "Risk",
        "Claude Verdict",
        "Shared Evidence",
        "Explanation",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    if not pairs:
        ws.cell(row=2, column=1, value="No pairs flagged — all essays below similarity thresholds.")
        _autosize(ws, max_width=60)
        return

    # Highest risk first so reviewers see the worst pairs at the top.
    risk_order = {"High": 0, "Medium": 1, "Low": 2}
    sorted_pairs = sorted(pairs, key=lambda p: risk_order.get(p.get("risk", ""), 3))

    for row_idx, p in enumerate(sorted_pairs, start=2):
        evidence = p.get("shared_evidence", []) or []
        values = [
            p.get("candidate_a", ""),
            p.get("candidate_b", ""),
            p.get("lexical_pct", ""),
            p.get("semantic_pct", ""),
            p.get("risk", ""),
            p.get("claude_verdict", ""),
            "\n".join(f"• {e}" for e in evidence),
            p.get("claude_explanation", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col >= 7:  # evidence + explanation are long-form text
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        risk_cell = ws.cell(row=row_idx, column=5)
        if p.get("risk") == "High":
            risk_cell.fill = red
        elif p.get("risk") == "Medium":
            risk_cell.fill = yellow

    _autosize(ws, max_width=60)


def _autosize(ws, max_width: int = 40) -> None:
    """Approximate column auto-sizing based on content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=0,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)
