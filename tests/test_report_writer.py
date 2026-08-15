"""Unit tests for the Detailed sheet's per-question summary columns.

Candidates answer the written task in whatever order they like, so the Q1/Q2/Q3
columns are filled by canonical question_number rather than by position in the
question_assessments list. These tests pin that behaviour, plus the fallbacks
for legacy results and error rows.

No API calls are made. Run from the project root with:

    venv\\Scripts\\python.exe -m unittest discover tests -v
"""

import datetime as dt
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from openpyxl import load_workbook

import report_writer as rw
import essay_grader as eg


# ------------------------------------------------------------
# Fixtures
# ------------------------------------------------------------
def _assessment(number, summary, key=""):
    return {
        "question_number": number,
        "question_key": key,
        "question_text": "",
        "summary": summary,
        "quality": "Good",
        "relevance": "High",
        "competencies_demonstrated": [],
    }


def _result(number, role, assessments):
    return {
        "candidate_number": number,
        "Role": role,
        "classification": "Priority Interview",
        "rationale": "",
        "cross_cutting_assessment": {},
        "strengths": [],
        "weaknesses": [],
        "writing_quality": {},
        "ai_usage_probability": "Low",
        "ai_usage_indicators": "",
        "question_assessments": assessments,
        "source_file": f"{number}_{role}_assignment.pdf",
    }


# ------------------------------------------------------------
# Column mapping
# ------------------------------------------------------------
class TestQuestionSummaries(unittest.TestCase):
    def test_out_of_sequence_answers_land_in_canonical_columns(self):
        """A TRI candidate who wrote 3, 1, 2 still lines up with everyone else."""
        summaries = rw._question_summaries([
            _assessment(3, "most challenging", "tri-most-challenging"),
            _assessment(1, "slats guidance", "tri-slats"),
            _assessment(2, "sbt design", "tri-sbt-behaviours"),
        ])
        self.assertEqual(summaries[1], "slats guidance")
        self.assertEqual(summaries[2], "sbt design")
        self.assertEqual(summaries[3], "most challenging")

    def test_two_question_role_leaves_q3_empty(self):
        """TFO has only two questions."""
        summaries = rw._question_summaries([
            _assessment(1, "ground school", "tfo-ground-school"),
            _assessment(2, "feedback", "tfo-feedback"),
        ])
        self.assertEqual(summaries.get(3, ""), "")

    def test_unnumbered_assessments_fall_back_to_position(self):
        """Legacy cached results carry no question_number."""
        summaries = rw._question_summaries([
            {"summary": "first"},
            {"summary": "second"},
            {"summary": "third"},
        ])
        self.assertEqual(summaries[1], "first")
        self.assertEqual(summaries[2], "second")
        self.assertEqual(summaries[3], "third")

    def test_string_question_numbers_are_accepted(self):
        summaries = rw._question_summaries([_assessment("2", "sbt design")])
        self.assertEqual(summaries[2], "sbt design")
        self.assertNotIn(1, summaries)

    def test_positional_fallback_never_overwrites_a_numbered_answer(self):
        summaries = rw._question_summaries([
            _assessment(1, "numbered slats"),
            {"summary": "unnumbered stray"},
        ])
        self.assertEqual(summaries[1], "numbered slats")

    def test_empty_and_malformed_input(self):
        self.assertEqual(rw._question_summaries([]), {})
        self.assertEqual(rw._question_summaries(["not a dict"]), {})

    def test_out_of_range_number_is_dropped_not_guessed_into_q1(self):
        """Better a blank cell than the wrong answer under the Q1 header."""
        self.assertEqual(rw._question_summaries([_assessment(9, "out of range")]), {})


# ------------------------------------------------------------
# End-to-end through the workbook
# ------------------------------------------------------------
class TestDetailedSheet(unittest.TestCase):
    def _write_and_read(self, results):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            rw.write_report(results, str(path))
            ws = load_workbook(path)["Detailed"]
            return [[c.value for c in row] for row in ws.iter_rows()]

    def test_headers_and_row_values(self):
        rows = self._write_and_read([
            _result("100001", "TRI", [
                _assessment(3, "most challenging"),
                _assessment(1, "slats guidance"),
                _assessment(2, "sbt design"),
            ]),
        ])
        headers = rows[0]
        self.assertEqual(headers[-3:], ["Q1 Summary", "Q2 Summary", "Q3 Summary"])
        self.assertEqual(rows[1][-3:], ["slats guidance", "sbt design", "most challenging"])

    def test_financial_year_is_on_the_detailed_sheet_too(self):
        """A Detailed row must be tie-able to a campaign on its own, or it
        cannot be joined once the List holds more than one year."""
        result = _result("100001", "TRI", [])
        result["campaign"] = "FY26"
        rows = self._write_and_read([result])
        self.assertEqual(rows[0][:2], ["Candidate Number", "Financial Year"])
        self.assertEqual(rows[1][1], 2026)

    def test_error_row_writes_blanks_rather_than_raising(self):
        error_row = eg._error_result("100002", "TRI", "Empty essay submission")
        rows = self._write_and_read([error_row])
        # openpyxl reads an empty-string cell back as None.
        self.assertEqual([v or "" for v in rows[1][-3:]], ["", "", ""])


# ------------------------------------------------------------
# Double Application column
# ------------------------------------------------------------
class TestDoubleApplication(unittest.TestCase):
    def test_same_number_under_two_roles_is_flagged(self):
        numbers = rw._double_application_numbers([
            {"candidate_number": "872524", "Role": "TRI"},
            {"candidate_number": "872524", "Role": "TFO TRI"},
            {"candidate_number": "111111", "Role": "LTC"},
        ])
        self.assertEqual(numbers, {"872524"})

    def test_single_role_candidate_is_not_flagged(self):
        self.assertEqual(
            rw._double_application_numbers([{"candidate_number": "1", "Role": "LTC"}]),
            set(),
        )

    def test_three_roles_still_flags_once(self):
        numbers = rw._double_application_numbers([
            {"candidate_number": "7", "Role": "LTC"},
            {"candidate_number": "7", "Role": "TRI"},
            {"candidate_number": "7", "Role": "TFO TRI"},
        ])
        self.assertEqual(numbers, {"7"})

    def test_rows_without_a_candidate_number_are_ignored(self):
        """Blank numbers must not group into a phantom double applicant."""
        numbers = rw._double_application_numbers([
            {"Role": "LTC"}, {"Role": "TRI"}, {"candidate_number": "", "Role": "TFO"},
        ])
        self.assertEqual(numbers, set())


class TestSimilarityFinancialYear(unittest.TestCase):
    """Pairs carry no campaign of their own; the sheet is stamped with the
    run's, which is sound because the screen never crosses campaigns."""

    def _similarity(self, campaign):
        pair = {
            "candidate_a": "111", "candidate_b": "222",
            "lexical_pct": 80.0, "semantic_pct": 90.0, "risk": "High",
            "claude_verdict": "copied", "shared_evidence": ["a passage"],
            "claude_explanation": "why",
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            rw.write_report(
                [_result("1", "TRI", [])], str(path),
                similarity_pairs=[pair], campaign=campaign,
            )
            ws = load_workbook(path)["Similarity"]
            return [[c.value for c in row] for row in ws.iter_rows()]

    def test_the_pair_row_carries_the_year(self):
        rows = self._similarity("FY26")
        self.assertEqual(rows[0][2], "Financial Year")
        self.assertEqual(rows[1][2], 2026)

    def test_risk_banding_survived_the_inserted_column(self):
        """Regression: Risk used to be located at a hard-coded column 5."""
        rows = self._similarity("FY26")
        self.assertEqual(rows[1][rows[0].index("Risk")], "High")

    def test_no_campaign_leaves_the_cell_blank(self):
        rows = self._similarity("")
        self.assertIsNone(rows[1][2])


class TestHistorySheet(unittest.TestCase):
    """The one sheet that crosses campaign boundaries."""

    def _history(self, rows):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            rw.write_report([_result("1", "TRI", [])], str(path), history=rows)
            ws = load_workbook(path)["History"]
            return [[c.value for c in row] for row in ws.iter_rows()]

    def _row(self, number, campaign, role="TRI", decision="NO", approval="REJECTED"):
        return {
            "candidate_number": number,
            "campaign": campaign,
            "role": role,
            "submitted": dt.date(2026, 7, 28),
            "interview_decision": decision,
            "final_approval": approval,
        }

    def test_a_returning_candidate_shows_every_campaign(self):
        rows = self._history([
            self._row("872524", "FY26", "TRI", "NO", "REJECTED"),
            self._row("872524", "FY27", "LTC", "PENDING", "PENDING"),
        ])
        self.assertEqual(
            rows[0],
            ["Candidate Number", "Financial Year", "Role", "Submitted",
             "Interview Decision", "Final Approval"],
        )
        # Shown as the number the Power Automate flows join on.
        self.assertEqual([r[1] for r in rows[1:]], [2026, 2027])
        # What actually happened last time — the question a re-applicant raises.
        self.assertEqual([r[4] for r in rows[1:]], ["NO", "PENDING"])
        self.assertEqual([r[5] for r in rows[1:]], ["REJECTED", "PENDING"])

    def test_a_single_campaign_candidate_is_omitted(self):
        """Their one row is already the Summary; repeating ~150 of them would
        bury the handful of returning candidates this sheet is for."""
        rows = self._history([
            self._row("872524", "FY26"),
            self._row("872524", "FY27"),
            self._row("111111", "FY26"),
        ])
        self.assertEqual({r[0] for r in rows[1:]}, {"872524"})

    def test_grouping_still_uses_the_campaign_not_the_displayed_year(self):
        """The number is presentation only. If the display change had leaked
        into the grouping, this candidate would stop being 'returning'."""
        rows = self._history([
            self._row("872524", "FY26"),
            self._row("872524", "FY27"),
        ])
        self.assertEqual([r[1] for r in rows[1:]], [2026, 2027])

    def test_nobody_returning_says_so_rather_than_looking_broken(self):
        rows = self._history([self._row("111111", "FY26")])
        self.assertIn("more than one campaign", str(rows[1][0]))

    def test_the_sheet_is_absent_when_no_history_is_passed(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            rw.write_report([_result("1", "TRI", [])], str(path))
            self.assertNotIn("History", load_workbook(path).sheetnames)


class TestSummarySheet(unittest.TestCase):
    def _summary(self, results):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"
            rw.write_report(results, str(path))
            ws = load_workbook(path)["Summary"]
            headers = [c.value for c in ws[1]]
            rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
            fills = {
                h: [ws.cell(row=r, column=headers.index(h) + 1).fill
                    for r in range(2, ws.max_row + 1)]
                for h in headers
            }
            return headers, rows, fills

    def test_column_reports_yes_and_no(self):
        headers, rows, _ = self._summary([
            _result("872524", "TRI", []),
            _result("872524", "TFO TRI", []),
            _result("111111", "LTC", []),
        ])
        col = headers.index("Double Application")
        self.assertEqual([r[col] for r in rows], ["Yes", "Yes", "No"])

    def test_campaign_and_submitted_are_written(self):
        row = _result("1", "TRI", [])
        row["campaign"] = "FY26"
        row["submitted"] = dt.date(2026, 7, 28)
        headers, rows, _ = self._summary([row])
        # The flows match on this, so it must be 2026 the number — not
        # "FY26", and not the string "2026".
        year = rows[0][headers.index("Financial Year")]
        self.assertEqual(year, 2026)
        self.assertIsInstance(year, int)
        # A real date, not text: Excel must sort it chronologically.
        self.assertEqual(rows[0][headers.index("Submitted")], dt.datetime(2026, 7, 28))

    def test_a_row_with_no_date_leaves_submitted_blank(self):
        """No recruitment list, or a candidate absent from it. Never a guess."""
        headers, rows, _ = self._summary([_result("1", "TRI", [])])
        self.assertIn(rows[0][headers.index("Submitted")], ("", None))

    def test_embargo_text_is_written_and_banded_red(self):
        flagged = _result("1", "TRI", [])
        flagged["embargo"] = "⚠ Re-applied 79d (2.6 months) after FY26 application"
        headers, rows, fills = self._summary([flagged, _result("2", "LTC", [])])
        col = headers.index("Embargo")
        self.assertTrue(str(rows[0][col]).startswith("⚠"))
        self.assertEqual(fills["Embargo"][0].start_color.rgb, "00FFC7CE")

    def test_an_unchecked_candidate_is_banded_but_not_condemned(self):
        """'?' must not look the same as a clean check, nor as a breach."""
        unknown = _result("1", "TRI", [])
        unknown["embargo"] = "? not in recruitment list"
        _, _, fills = self._summary([unknown])
        self.assertEqual(fills["Embargo"][0].start_color.rgb, "00FFEB9C")

    def test_a_clear_candidate_leaves_the_cell_blank_and_unfilled(self):
        _, rows, fills = self._summary([_result("1", "TRI", [])])
        self.assertIn(rows[0][3], ("", None))
        self.assertNotEqual(fills["Embargo"][0].start_color.rgb, "00FFC7CE")

    def test_classification_band_still_lands_on_the_right_cell(self):
        """Regression: the colour band used to assume Classification was 3rd.

        Inserting Double Application ahead of it would silently colour the
        wrong column.
        """
        headers, rows, fills = self._summary([_result("1", "LTC", [])])
        self.assertGreater(
            headers.index("Classification"), 2,
            "Classification is no longer 3rd — this test guards that move",
        )
        green = "00C6EFCE"
        self.assertEqual(fills["Classification"][0].start_color.rgb, green)
        # And the column it displaced must NOT have been coloured.
        self.assertNotEqual(
            fills["Double Application"][0].start_color.rgb, green
        )


# ------------------------------------------------------------
# Grader-side normalisation
# ------------------------------------------------------------
class TestNormaliseQuestionAssessments(unittest.TestCase):
    def test_sorts_into_canonical_order_and_coerces_numbers(self):
        normalised = eg._normalise_question_assessments([
            _assessment("3", "third"),
            _assessment(1.0, "first"),
            _assessment(2, "second"),
        ])
        self.assertEqual([a["question_number"] for a in normalised], [1, 2, 3])
        self.assertEqual([a["summary"] for a in normalised], ["first", "second", "third"])

    def test_unnumbered_entries_sort_last_in_original_order(self):
        normalised = eg._normalise_question_assessments([
            {"summary": "stray a"},
            _assessment(1, "numbered"),
            {"summary": "stray b"},
        ])
        self.assertEqual(
            [a["summary"] for a in normalised],
            ["numbered", "stray a", "stray b"],
        )
        self.assertIsNone(normalised[1]["question_number"])

    def test_non_list_and_non_dict_input_is_dropped(self):
        self.assertEqual(eg._normalise_question_assessments(None), [])
        self.assertEqual(eg._normalise_question_assessments("nope"), [])
        self.assertEqual(eg._normalise_question_assessments(["nope"]), [])

    def test_normalise_result_applies_it(self):
        result = eg._normalise_result(
            {"question_assessments": [_assessment(2, "second"), _assessment(1, "first")]},
            "100003",
            "TRI",
        )
        self.assertEqual(
            [a["question_number"] for a in result["question_assessments"]], [1, 2]
        )

    def test_does_not_mutate_the_caller_s_dicts(self):
        original = [{"question_number": "1", "summary": "first"}]
        eg._normalise_question_assessments(original)
        self.assertEqual(original[0]["question_number"], "1")


if __name__ == "__main__":
    unittest.main()
