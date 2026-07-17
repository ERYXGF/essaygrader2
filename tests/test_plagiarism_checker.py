"""Unit tests for the plagiarism screening feature.

No API calls are made: Claude pair verdicts are mocked. Run from the project
root with:

    venv\\Scripts\\python.exe -m unittest discover tests -v
"""

import sys
import unittest
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import plagiarism_checker as pc
from report_writer import write_report

from openpyxl import load_workbook


# ------------------------------------------------------------
# Test fixtures
# ------------------------------------------------------------
ESSAY_ORIGINAL = (
    "To develop situation awareness in the simulator I would design a scenario "
    "arriving into Lyon with CAT3 weather and introduce a NAV ILS fault. I would "
    "freeze the simulator and use facilitation to ensure the crew understands the "
    "degraded landing capability, then let them apply their decision making model "
    "and divert to Geneva where the weather is currently better than CAT2 minima."
)

# Near-verbatim copy with light edits (a few words changed).
ESSAY_COPIED = (
    "To develop situation awareness in the simulator I would design a scenario "
    "arriving into Lyon with CAT3 weather and introduce a NAV ILS failure. I would "
    "pause the simulator and use facilitation to ensure the crew understands the "
    "degraded landing capability, then let them apply their decision making model "
    "and divert to Geneva where the weather is currently better than CAT2 minima."
)

ESSAY_UNRELATED = (
    "Ground school can feel repetitive for experienced crew, so my strategy centres "
    "on adult learning principles. I open every module with a real operational event "
    "from the previous season, ask the group to diagnose it before revealing the "
    "outcome, and close with an anonymous exit poll measuring what participants will "
    "change in their next duty. Feedback loops keep the content honest and relevant."
)


def _essay(number, text, role="TRI"):
    return {
        "candidate_number": number,
        "role": role,
        "essay_text": text,
        "source_file": f"{number}_{role}_assignment.pdf",
    }


def _fake_verdict(verdict="confirmed", explanation="Near-verbatim overlap.",
                  evidence=("shared passage",)):
    return {
        "verdict": verdict,
        "explanation": explanation,
        "shared_evidence": list(evidence),
    }


# ------------------------------------------------------------
# Lexical metrics
# ------------------------------------------------------------
class TestLexicalMetrics(unittest.TestCase):
    def test_identical_texts_score_maximum(self):
        tokens = pc._tokenize(ESSAY_ORIGINAL)
        shingles = pc._shingles(tokens, pc.SHINGLE_SIZE)
        self.assertEqual(pc._jaccard(shingles, shingles), 1.0)

        vecs = pc._tfidf_vectors([tokens, tokens])
        self.assertAlmostEqual(pc._cosine(vecs[0], vecs[1]), 1.0, places=9)

    def test_copied_pair_scores_high(self):
        tok_a = pc._tokenize(ESSAY_ORIGINAL)
        tok_b = pc._tokenize(ESSAY_COPIED)
        jac = pc._jaccard(
            pc._shingles(tok_a, pc.SHINGLE_SIZE),
            pc._shingles(tok_b, pc.SHINGLE_SIZE),
        )
        self.assertGreater(jac, pc.LEXICAL_HIGH_THRESHOLD)

        vecs = pc._tfidf_vectors([tok_a, tok_b])
        self.assertGreater(pc._cosine(vecs[0], vecs[1]), pc.SEMANTIC_SCREEN_THRESHOLD)

    def test_unrelated_pair_scores_low(self):
        tok_a = pc._tokenize(ESSAY_ORIGINAL)
        tok_b = pc._tokenize(ESSAY_UNRELATED)
        jac = pc._jaccard(
            pc._shingles(tok_a, pc.SHINGLE_SIZE),
            pc._shingles(tok_b, pc.SHINGLE_SIZE),
        )
        self.assertLess(jac, pc.LEXICAL_SCREEN_THRESHOLD)

    def test_tokenizer_ignores_case_and_punctuation(self):
        self.assertEqual(
            pc._tokenize("Freeze, the SIM!"), pc._tokenize("freeze the sim")
        )

    def test_empty_inputs_are_safe(self):
        self.assertEqual(pc._jaccard(frozenset(), frozenset({"a b c"})), 0.0)
        self.assertEqual(pc._cosine({}, {"a": 1.0}), 0.0)


# ------------------------------------------------------------
# Screening / check_plagiarism (Claude mocked)
# ------------------------------------------------------------
class TestCheckPlagiarism(unittest.TestCase):
    def test_copied_pair_is_flagged_and_canonically_ordered(self):
        # Higher number listed first on purpose — output must be reordered.
        essays = [
            _essay("677667", ESSAY_COPIED),
            _essay("222333", ESSAY_ORIGINAL),
            _essay("98765", ESSAY_UNRELATED, role="LTC"),
        ]
        with patch.object(pc, "_claude_pair_verdict", return_value=_fake_verdict()), \
             patch.object(pc, "_build_client", return_value=object()), \
             patch.object(pc, "_load_pair_review_prompt", return_value="prompt"):
            pairs = pc.check_plagiarism(essays)

        self.assertEqual(len(pairs), 1)
        pair = pairs[0]
        self.assertEqual(pair["candidate_a"], "222333")  # lower number first
        self.assertEqual(pair["candidate_b"], "677667")
        self.assertEqual(pair["risk"], "High")
        self.assertGreater(pair["lexical_pct"], 20)
        self.assertGreater(pair["semantic_pct"], 50)
        self.assertEqual(pair["claude_verdict"], "confirmed")

    def test_clean_batch_returns_no_pairs_and_never_builds_client(self):
        essays = [
            _essay("111", ESSAY_ORIGINAL),
            _essay("222", ESSAY_UNRELATED),
        ]
        with patch.object(pc, "_build_client",
                          side_effect=AssertionError("client built for clean batch")):
            self.assertEqual(pc.check_plagiarism(essays), [])

    def test_empty_essays_are_skipped(self):
        essays = [
            _essay("111", ""),
            _essay("222", "   "),
            _essay("333", ESSAY_ORIGINAL),
        ]
        with patch.object(pc, "_build_client",
                          side_effect=AssertionError("client built unexpectedly")):
            self.assertEqual(pc.check_plagiarism(essays), [])

    def test_extreme_verbatim_overlap_is_high_even_if_claude_equivocates(self):
        essays = [
            _essay("111", ESSAY_ORIGINAL),
            _essay("222", ESSAY_ORIGINAL),
        ]
        verdict = _fake_verdict(verdict="likely_coincidence", evidence=())
        with patch.object(pc, "_claude_pair_verdict", return_value=verdict), \
             patch.object(pc, "_build_client", return_value=object()), \
             patch.object(pc, "_load_pair_review_prompt", return_value="prompt"):
            pairs = pc.check_plagiarism(essays)
        self.assertEqual(pairs[0]["risk"], "High")


# ------------------------------------------------------------
# Classification overrides + Summary flags
# ------------------------------------------------------------
class TestApplyOverrides(unittest.TestCase):
    def _pair(self, risk="High", a="111", b="222", verdict="confirmed"):
        return {
            "candidate_a": a,
            "candidate_b": b,
            "lexical_pct": 87.0,
            "semantic_pct": 92.0,
            "risk": risk,
            "claude_verdict": verdict,
            "claude_explanation": "Near-verbatim overlap.",
            "shared_evidence": ["shared passage"],
        }

    def _result(self, number, classification, rationale="Strong answer."):
        return {
            "candidate_number": number,
            "classification": classification,
            "rationale": rationale,
        }

    def test_high_risk_downgrades_priority_interview_to_maybe(self):
        results = [
            self._result("111", "Priority Interview"),
            self._result("222", "Borderline — Human Review Required"),
        ]
        pc.apply_plagiarism_overrides(results, [self._pair()])

        for r in results:
            self.assertEqual(r["classification"], "Maybe")
            self.assertIn("[Plagiarism override]", r["rationale"])
            self.assertIn("⚠ HIGH", r["plagiarism_flag"])

        # Both rows must name their partner, not themselves.
        self.assertIn("222", results[0]["plagiarism_flag"])
        self.assertIn("111", results[1]["plagiarism_flag"])

    def test_do_not_interview_is_never_upgraded(self):
        results = [self._result("111", "Do Not Interview"),
                   self._result("222", "Priority Interview")]
        pc.apply_plagiarism_overrides(results, [self._pair()])
        self.assertEqual(results[0]["classification"], "Do Not Interview")
        self.assertIn("⚠ HIGH", results[0]["plagiarism_flag"])  # still flagged

    def test_medium_risk_flags_without_downgrading(self):
        results = [self._result("111", "Priority Interview"),
                   self._result("222", "Maybe")]
        pc.apply_plagiarism_overrides(
            results, [self._pair(risk="Medium", verdict="suspicious")]
        )
        self.assertEqual(results[0]["classification"], "Priority Interview")
        self.assertIn("⚠ MEDIUM", results[0]["plagiarism_flag"])
        self.assertNotIn("[Plagiarism override]", results[0]["rationale"])

    def test_low_risk_pairs_produce_no_summary_flag(self):
        results = [self._result("111", "Priority Interview"),
                   self._result("222", "Maybe")]
        pc.apply_plagiarism_overrides(
            results, [self._pair(risk="Low", verdict="likely_coincidence")]
        )
        self.assertEqual(results[0]["plagiarism_flag"], "")
        self.assertEqual(results[0]["classification"], "Priority Interview")

    def test_clean_candidates_get_empty_flag_and_error_rows_not_screened(self):
        results = [
            self._result("333", "Maybe"),
            self._result("444", "error"),
        ]
        pc.apply_plagiarism_overrides(results, [])
        self.assertEqual(results[0]["plagiarism_flag"], "")
        self.assertEqual(results[1]["plagiarism_flag"], pc.NOT_SCREENED_FLAG)


# ------------------------------------------------------------
# Excel report integration
# ------------------------------------------------------------
class TestReportWriter(unittest.TestCase):
    def _base_result(self, number, classification, flag=""):
        return {
            "candidate_number": number,
            "Role": "TRI",
            "classification": classification,
            "rationale": "r",
            "cross_cutting_assessment": {},
            "strengths": [],
            "weaknesses": [],
            "writing_quality": {},
            "ai_usage_probability": "Low",
            "ai_usage_indicators": "",
            "question_assessments": [],
            "source_file": f"{number}_TRI_assignment.pdf",
            "plagiarism_flag": flag,
        }

    def test_report_has_flag_column_and_similarity_sheet(self):
        import tempfile, os
        results = [
            self._base_result("111", "Maybe",
                              flag="⚠ HIGH: lexical 87% / semantic 92% with 222 — Claude: confirmed"),
            self._base_result("222", "Maybe",
                              flag="⚠ HIGH: lexical 87% / semantic 92% with 111 — Claude: confirmed"),
            self._base_result("333", "Priority Interview"),
        ]
        pairs = [{
            "candidate_a": "111", "candidate_b": "222",
            "lexical_pct": 87.0, "semantic_pct": 92.0,
            "risk": "High", "claude_verdict": "confirmed",
            "claude_explanation": "Near-verbatim overlap.",
            "shared_evidence": ["shared passage"],
        }]

        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            write_report(results=results, output_path=path, similarity_pairs=pairs)
            wb = load_workbook(path)

            self.assertEqual(wb.sheetnames, ["Summary", "Detailed", "Similarity"])

            ws1 = wb["Summary"]
            headers = [c.value for c in ws1[1]]
            self.assertEqual(headers[-1], "Plagiarism Flag")  # last column
            flag_col = len(headers)
            self.assertIn("⚠ HIGH", ws1.cell(row=2, column=flag_col).value)
            self.assertIsNone(ws1.cell(row=4, column=flag_col).value)  # clean row empty

            ws3 = wb["Similarity"]
            self.assertEqual(ws3.cell(row=2, column=1).value, "111")
            self.assertEqual(ws3.cell(row=2, column=5).value, "High")
            self.assertIn("shared passage", ws3.cell(row=2, column=7).value)

    def test_report_without_pairs_is_backward_compatible(self):
        import tempfile, os
        results = [self._base_result("111", "Maybe")]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            write_report(results=results, output_path=path)  # old signature
            wb = load_workbook(path)
            self.assertEqual(wb.sheetnames, ["Summary", "Detailed"])

    def test_report_with_empty_pairs_writes_clean_similarity_sheet(self):
        import tempfile, os
        results = [self._base_result("111", "Maybe")]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "report.xlsx")
            write_report(results=results, output_path=path, similarity_pairs=[])
            wb = load_workbook(path)
            self.assertIn("Similarity", wb.sheetnames)
            self.assertIn("No pairs flagged", wb["Similarity"].cell(row=2, column=1).value)


if __name__ == "__main__":
    unittest.main()
